#!/usr/bin/env python3
"""
AI Trader — Build Dashboard
Regenerate index.html from latest data in Excel + JSON data
"""

import argparse
import json
import sys
from datetime import datetime
import openpyxl
import warnings
warnings.filterwarnings('ignore')

TODAY_SHORT = datetime.now().strftime("%-d %b")


def advice_badge(advice_text: str) -> str:
    """Convert AI Advice text to badge class"""
    if not advice_text:
        return ("advice-hold", "Hold")
    t = advice_text.lower()
    if "cut loss" in t or "cut all" in t:
        return ("advice-cutloss", "⛔ Cut Loss")
    elif "reduce" in t:
        return ("advice-reduce", "⚠️ Reduce")
    elif "hold / add" in t or "increase" in t:
        return ("advice-increase", "✅ Hold/Add")
    elif "long-term" in t:
        return ("advice-longterm", "🔷 Long-term")
    else:
        return ("advice-hold", "🔵 Hold")


def read_portfolio(ws, stocks_data: dict, usdthb: float):
    """Read portfolio data from Excel sheet"""
    import re
    VALID_TICKER = re.compile(r'^[A-Z]{1,6}$')
    rows = []
    for row_idx in range(5, 70):
        num    = ws.cell(row=row_idx, column=1).value
        sym    = ws.cell(row=row_idx, column=2).value
        cost   = ws.cell(row=row_idx, column=3).value
        price  = ws.cell(row=row_idx, column=4).value
        val_usd= ws.cell(row=row_idx, column=7).value
        val_thb= ws.cell(row=row_idx, column=8).value
        advice = ws.cell(row=row_idx, column=9).value
        updated= ws.cell(row=row_idx, column=11).value

        # Stop at first empty cell — marks end of stock table
        if sym is None:
            break

        if not isinstance(sym, str):
            continue

        sym = sym.strip()

        # Only accept valid ticker symbols (1-6 uppercase letters)
        if not VALID_TICKER.match(sym):
            continue

        # Use price from JSON if available
        if sym in stocks_data:
            price = stocks_data[sym]["price"]

        # Calculate values
        cost_f   = float(cost) if isinstance(cost, (int, float)) else 0
        price_f  = float(price) if isinstance(price, (int, float)) else 0
        val_usd_f= float(val_usd) if isinstance(val_usd, (int, float)) else 0
        val_thb_f= val_usd_f * usdthb

        if cost_f > 0 and price_f > 0 and val_usd_f > 0:
            qty = val_usd_f / price_f if price_f > 0 else 0
            pnl_usd = val_usd_f * (price_f - cost_f) / price_f
            pnl_pct = (price_f - cost_f) / cost_f * 100
        else:
            pnl_usd = 0
            pnl_pct = 0

        rows.append({
            "num": num,
            "symbol": sym,
            "cost": cost_f,
            "price": price_f,
            "pnl_usd": round(pnl_usd, 2),
            "pnl_pct": round(pnl_pct, 2),
            "val_usd": round(val_usd_f, 2),
            "val_thb": round(val_thb_f, 0),
            "advice": str(advice) if advice else "",
            "updated": str(updated) if updated else "—",
        })
    return rows


def build_html(portfolio: list, stocks_data: dict, trade_data: dict,
               usdthb: float, updated_at: str, default_tab: str = "portfolio") -> str:

    total_val_usd = sum(r["val_usd"] for r in portfolio)
    total_val_thb = total_val_usd * usdthb
    total_pnl_usd = sum(r["pnl_usd"] for r in portfolio)
    total_pnl_pct = (total_pnl_usd / (total_val_usd - total_pnl_usd) * 100) if (total_val_usd - total_pnl_usd) > 0 else 0
    gainers = [r for r in portfolio if r["pnl_pct"] > 0]
    losers  = [r for r in portfolio if r["pnl_pct"] < 0]

    # ──── Portfolio Table Rows ────
    table_rows_html = ""
    for r in portfolio:
        badge_class, badge_text = advice_badge(r["advice"])
        pnl_class = "positive" if r["pnl_usd"] >= 0 else "negative"
        pct_class  = "positive" if r["pnl_pct"] >= 0 else "negative"
        sign       = "+" if r["pnl_usd"] >= 0 else ""
        pct_sign   = "+" if r["pnl_pct"] >= 0 else ""
        port_pct   = round(r["val_usd"] / total_val_usd * 100, 1) if total_val_usd > 0 else 0
        table_rows_html += f"""
      <tr data-advice="{badge_class}">
        <td style="color:var(--muted)">{r['num']}</td>
        <td><span class="ticker">{r['symbol']}</span></td>
        <td>${r['cost']:.4f}</td>
        <td>${r['price']:.4f}</td>
        <td class="{pnl_class}">{sign}${abs(r['pnl_usd']):.2f}</td>
        <td class="{pct_class}">{pct_sign}{r['pnl_pct']:.2f}%</td>
        <td>${r['val_usd']:.2f}</td>
        <td>฿{r['val_thb']:,.0f}</td>
        <td>{port_pct:.1f}%</td>
        <td><span class="advice-badge {badge_class}">{badge_text}</span></td>
        <td style="color:var(--muted);font-size:12px">{r['updated']}</td>
      </tr>"""

    # ──── Donut Chart Data ────
    donut_labels = json.dumps([r["symbol"] for r in portfolio])
    donut_data   = json.dumps([r["val_usd"] for r in portfolio])
    donut_colors = json.dumps([
        "#58a6ff","#3fb950","#d29922","#f85149","#bc8cff","#ffa657",
        "#39d353","#1f6feb","#db6d28","#388bfd","#56d364","#e3b341",
        "#ff7b72","#d2a8ff","#79c0ff","#a5d6ff","#7ee787","#ffa198",
        "#cae8ff","#ffd700"
    ][:len(portfolio)])

    # ──── P&L Chart Data ────
    pl_labels = json.dumps([r["symbol"] for r in portfolio])
    pl_data   = json.dumps([r["pnl_usd"] for r in portfolio])
    pl_colors = json.dumps(["#3fb95099" if r["pnl_usd"] >= 0 else "#f8514999" for r in portfolio])
    pl_borders= json.dumps(["#3fb950" if r["pnl_usd"] >= 0 else "#f85149" for r in portfolio])

    # ──── Position Chart Data ────
    pos_labels = json.dumps([r["symbol"] for r in portfolio])
    pos_data   = json.dumps([r["val_thb"] for r in portfolio])

    # ──── AI Trading Section — multi-stock ────
    week_trades = trade_data.get("week_trades", [])
    budget_total = week_trades[0]["budget_total"] if week_trades else 200

    def rsi_color_fn(rsi):
        if rsi >= 70: return "var(--red)"
        if rsi <= 30: return "var(--blue)"
        if rsi >= 55: return "var(--green)"
        if rsi <= 45: return "var(--yellow)"
        return "var(--text)"

    def build_trade_card(t: dict, card_idx: int) -> str:
        sym      = t["sym"]
        company  = t["company"]
        price    = t["price"]
        entry    = t["entry"]
        sl       = t["sl"]
        sl_pct   = t["sl_pct"]
        tp1      = t["tp1"]
        tp2      = t["tp2"]
        tp3      = t["tp3"]
        rr       = t["rr"]
        strategy = t["strategy"]
        invest   = t["invest"]
        rsi1d    = t["rsi1d"]
        rsi1h    = t["rsi1h"]
        pe       = t["pe"]

        # Compute TP % from entry
        tp1_pct = round((tp1 - entry) / entry * 100, 1) if entry > 0 and tp1 > 0 else 5
        tp2_pct = round((tp2 - entry) / entry * 100, 1) if entry > 0 and tp2 > 0 else 8
        tp3_pct = round((tp3 - entry) / entry * 100, 1) if entry > 0 and tp3 > 0 else 12
        sl_pct_show = round(sl_pct * 100, 1) if sl_pct and abs(sl_pct) < 1 else round(sl_pct, 1) if sl_pct else -5
        if sl_pct_show > 0: sl_pct_show = -sl_pct_show

        # Price vs entry badge
        if price > 0 and entry > 0:
            diff_pct = (price - entry) / entry * 100
            price_badge = f"({'▲' if diff_pct >= 0 else '▼'}{abs(diff_pct):.1f}% vs entry)"
            price_color = "#3fb950" if diff_pct >= 0 else "#f85149"
        else:
            price_badge = ""
            price_color = "var(--text)"

        # Scenario probabilities
        if rsi1d < 40:
            bull_pct, neutral_pct, bear_pct = 50, 30, 20
        elif rsi1d > 65:
            bull_pct, neutral_pct, bear_pct = 30, 35, 35
        else:
            bull_pct, neutral_pct, bear_pct = 45, 35, 20
        ref = price if price > 0 else entry
        bull_range    = f"${ref*1.05:.0f}–${ref*1.12:.0f}"
        neutral_range = f"${ref*0.97:.0f}–${ref*1.03:.0f}"
        bear_range    = f"below ${ref*0.95:.0f}"

        rsi1d_color = rsi_color_fn(rsi1d)
        rsi1h_color = rsi_color_fn(rsi1h)
        rsi1d_desc  = 'Overbought ⚠️' if rsi1d >= 70 else ('Oversold ⚡' if rsi1d <= 30 else 'Neutral ✅')
        rsi1h_desc  = 'Overbought ⚠️' if rsi1h >= 70 else ('Oversold ⚡' if rsi1h <= 30 else 'Neutral ✅')

        return f"""
  <div class="trade-card" style="margin-bottom:28px;">
    <div style="margin-bottom:16px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px;">
      <div>
        <div class="trade-ticker">{sym} <span style="font-size:15px;color:var(--muted)">— {company}</span></div>
        <div style="font-size:13px;color:var(--muted);margin-top:4px;">
          P/E: {pe} &nbsp;|&nbsp; Budget: ${invest:,.0f} &nbsp;|&nbsp; 📅 {updated_at}
          <span class="updated-badge">🤖 AI Auto</span>
        </div>
      </div>
      <div style="text-align:right;">
        <div style="font-size:22px;font-weight:700;color:{price_color}">${price:.2f}</div>
        <div style="font-size:12px;color:var(--muted)">{price_badge}</div>
      </div>
    </div>

    <!-- PRICE LEVELS -->
    <div class="section-title"><div class="dot"></div> Entry / TP / SL</div>
    <div class="levels-grid">
      <div class="level-item entry">
        <div class="level-label">🎯 Entry</div>
        <div class="level-value blue-text">${entry:.2f}</div>
        <div class="level-sub">Target entry</div>
      </div>
      <div class="level-item tp1">
        <div class="level-label">✅ TP1 +{tp1_pct:.1f}%</div>
        <div class="level-value positive">${tp1:.2f}</div>
        <div class="level-sub">1st target</div>
      </div>
      <div class="level-item tp2">
        <div class="level-label">✅ TP2 +{tp2_pct:.1f}%</div>
        <div class="level-value positive">${tp2:.2f}</div>
        <div class="level-sub">2nd target</div>
      </div>
      <div class="level-item tp3">
        <div class="level-label">✅ TP3 +{tp3_pct:.1f}%</div>
        <div class="level-value positive">${tp3:.2f}</div>
        <div class="level-sub">Max target</div>
      </div>
      <div class="level-item sl">
        <div class="level-label">🛑 SL {sl_pct_show:.1f}%</div>
        <div class="level-value negative">${sl:.2f}</div>
        <div class="level-sub">Cut all</div>
      </div>
    </div>

    <!-- SCENARIOS -->
    <div class="section-title" style="margin-top:16px;"><div class="dot"></div> Scenario Analysis</div>
    <div class="scenarios">
      <div class="scenario bull-scenario">
        <div class="scenario-icon">🐂</div>
        <div class="scenario-name">Bullish</div>
        <div class="scenario-pct">{bull_pct}%</div>
        <div class="scenario-range">Target {bull_range}</div>
        <div class="prob-bar-wrap"><div class="prob-bar bull-bar" style="width:{bull_pct}%"></div></div>
      </div>
      <div class="scenario neutral-scenario">
        <div class="scenario-icon">📊</div>
        <div class="scenario-name">Neutral / Sideways</div>
        <div class="scenario-pct">{neutral_pct}%</div>
        <div class="scenario-range">{neutral_range}</div>
        <div class="prob-bar-wrap"><div class="prob-bar neutral-bar" style="width:{neutral_pct}%"></div></div>
      </div>
      <div class="scenario bear-scenario">
        <div class="scenario-icon">🐻</div>
        <div class="scenario-name">Bearish</div>
        <div class="scenario-pct">{bear_pct}%</div>
        <div class="scenario-range">{bear_range}</div>
        <div class="prob-bar-wrap"><div class="prob-bar bear-bar" style="width:{bear_pct}%"></div></div>
      </div>
    </div>

    <!-- RSI -->
    <div class="section-title" style="margin-top:16px;"><div class="dot"></div> RSI Indicators</div>
    <div class="rsi-grid">
      <div class="rsi-item">
        <div class="rsi-label">RSI (1D — Daily)</div>
        <div class="rsi-bar-wrap"><div class="rsi-marker" style="left:{min(rsi1d,99)}%"></div></div>
        <div class="rsi-value" style="color:{rsi1d_color}">{rsi1d}</div>
        <div class="rsi-desc">{rsi1d_desc}</div>
      </div>
      <div class="rsi-item">
        <div class="rsi-label">RSI (1H — Hourly)</div>
        <div class="rsi-bar-wrap"><div class="rsi-marker" style="left:{min(rsi1h,99)}%"></div></div>
        <div class="rsi-value" style="color:{rsi1h_color}">{rsi1h}</div>
        <div class="rsi-desc">{rsi1h_desc}</div>
      </div>
    </div>

    <!-- STRATEGY -->
    <div class="strategy-box" style="margin-top:16px;">
      <h4>⚡ Strategy — R/R {rr}</h4>
      <p style="white-space:pre-line">{strategy}</p>
    </div>
  </div>"""

    # Build all trade cards
    trade_cards_html = "".join(build_trade_card(t, i) for i, t in enumerate(week_trades))

    # Summary bar for AI tab
    syms_display = " / ".join(t["sym"] for t in week_trades) if week_trades else "N/A"
    total_invest = sum(t["invest"] for t in week_trades)

    def rsi_color(rsi):
        if rsi >= 70: return "var(--red)"
        if rsi <= 30: return "var(--blue)"
        if rsi >= 55: return "var(--green)"
        if rsi <= 45: return "var(--yellow)"
        return "var(--text)"

    def ema_row(label, val, price, timeframe):
        pos = "above" if price > val else "below"
        arrow = "↑ Above" if price > val else "↓ Below"
        color = "var(--red)" if price > val else "var(--green)"
        return f"""<tr>
          <td>EMA{label} ({timeframe})</td>
          <td class="{pos}">${val:.2f}</td>
          <td style="color:{color}">{arrow}</td>
        </tr>"""

    pnl_class_total = "positive" if total_pnl_usd >= 0 else "negative"
    pnl_sign_total  = "+" if total_pnl_usd >= 0 else ""

    # ──── Full HTML ────
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>📊 Stock Portfolio Dashboard</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<style>
  :root {{
    --bg: #0d1117; --card: #161b22; --card2: #1c2128; --border: #30363d;
    --text: #e6edf3; --muted: #8b949e; --green: #3fb950; --red: #f85149;
    --yellow: #d29922; --blue: #58a6ff; --purple: #bc8cff; --orange: #ffa657;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ background: var(--bg); color: var(--text); font-family: 'Segoe UI', system-ui, sans-serif; min-height: 100vh; }}
  .header {{ background: linear-gradient(135deg,#0d1117,#161b22,#1c2128); border-bottom: 1px solid var(--border); padding: 24px 32px; display: flex; align-items: center; gap: 16px; }}
  .header-icon {{ font-size: 40px; }}
  .header h1 {{ font-size: 24px; font-weight: 700; color: var(--blue); }}
  .header p {{ font-size: 13px; color: var(--muted); margin-top: 4px; }}
  .badge {{ background: #1f6feb22; border: 1px solid #1f6feb66; color: var(--blue); padding: 3px 10px; border-radius: 20px; font-size: 12px; font-weight: 600; margin-left: auto; }}
  .tabs {{ display: flex; border-bottom: 1px solid var(--border); padding: 0 32px; background: var(--card); }}
  .tab {{ padding: 14px 24px; cursor: pointer; font-size: 14px; font-weight: 600; color: var(--muted); border-bottom: 2px solid transparent; transition: all 0.2s; }}
  .tab:hover {{ color: var(--text); }}
  .tab.active {{ color: var(--blue); border-bottom-color: var(--blue); }}
  .main {{ padding: 24px 32px; max-width: 1600px; margin: 0 auto; }}
  .summary-grid {{ display: grid; grid-template-columns: repeat(auto-fit,minmax(200px,1fr)); gap: 16px; margin-bottom: 24px; }}
  .card {{ background: var(--card); border: 1px solid var(--border); border-radius: 12px; padding: 20px; transition: border-color 0.2s; }}
  .card:hover {{ border-color: #58a6ff44; }}
  .card-label {{ font-size: 12px; color: var(--muted); font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 8px; }}
  .card-value {{ font-size: 26px; font-weight: 700; }}
  .card-sub {{ font-size: 13px; color: var(--muted); margin-top: 4px; }}
  .positive {{ color: var(--green); }} .negative {{ color: var(--red); }}
  .neutral {{ color: var(--yellow); }} .blue-text {{ color: var(--blue); }}
  .section-title {{ font-size: 18px; font-weight: 700; margin-bottom: 16px; display: flex; align-items: center; gap: 10px; }}
  .section-title .dot {{ width: 8px; height: 8px; border-radius: 50%; background: var(--blue); }}
  .table-wrap {{ overflow-x: auto; border-radius: 12px; border: 1px solid var(--border); margin-bottom: 32px; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  thead th {{ background: var(--card2); padding: 12px 14px; text-align: left; font-weight: 600; color: var(--muted); font-size: 11px; text-transform: uppercase; letter-spacing: 0.5px; border-bottom: 1px solid var(--border); white-space: nowrap; }}
  tbody tr {{ border-bottom: 1px solid #21262d; transition: background 0.15s; }}
  tbody tr:hover {{ background: #1c2128; }}
  tbody tr:last-child {{ border-bottom: none; }}
  td {{ padding: 12px 14px; vertical-align: middle; }}
  .ticker {{ font-weight: 700; font-size: 14px; color: var(--blue); }}
  .advice-badge {{ display: inline-block; padding: 3px 8px; border-radius: 6px; font-size: 11px; font-weight: 600; white-space: nowrap; }}
  .advice-hold {{ background: #1f6feb22; color: var(--blue); border: 1px solid #1f6feb44; }}
  .advice-increase {{ background: #1a7f3722; color: var(--green); border: 1px solid #1a7f3744; }}
  .advice-reduce {{ background: #9e6a0322; color: var(--yellow); border: 1px solid #9e6a0344; }}
  .advice-cutloss {{ background: #da363322; color: var(--red); border: 1px solid #da363344; }}
  .advice-longterm {{ background: #bc8cff22; color: var(--purple); border: 1px solid #bc8cff44; }}
  .charts-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin-bottom: 32px; }}
  .chart-card {{ background: var(--card); border: 1px solid var(--border); border-radius: 12px; padding: 20px; }}
  .chart-card h3 {{ font-size: 15px; font-weight: 600; margin-bottom: 16px; color: var(--text); }}
  .chart-container {{ position: relative; height: 280px; }}
  .filter-bar {{ display: flex; gap: 8px; flex-wrap: wrap; margin-bottom: 16px; }}
  .filter-btn {{ padding: 6px 14px; border-radius: 20px; border: 1px solid var(--border); background: var(--card2); color: var(--muted); font-size: 12px; font-weight: 600; cursor: pointer; transition: all 0.2s; }}
  .filter-btn:hover,.filter-btn.active {{ background: #1f6feb33; border-color: var(--blue); color: var(--blue); }}
  .section {{ display: none; }} .section.active {{ display: block; }}
  .trade-card {{ background: var(--card); border: 1px solid var(--border); border-radius: 12px; padding: 24px; margin-bottom: 20px; }}
  .trade-ticker {{ font-size: 36px; font-weight: 800; color: var(--blue); }}
  .levels-grid {{ display: grid; grid-template-columns: repeat(auto-fit,minmax(160px,1fr)); gap: 12px; margin: 16px 0; }}
  .level-item {{ background: var(--card2); border-radius: 8px; padding: 14px; border-left: 3px solid; }}
  .level-label {{ font-size: 11px; color: var(--muted); font-weight: 600; margin-bottom: 4px; }}
  .level-value {{ font-size: 20px; font-weight: 700; }}
  .level-sub {{ font-size: 12px; color: var(--muted); margin-top: 2px; }}
  .entry {{ border-color: var(--blue); }} .tp1 {{ border-color: #3fb95099; }}
  .tp2 {{ border-color: var(--green); }} .tp3 {{ border-color: #2ea04399; }}
  .sl {{ border-color: var(--red); }}
  .scenarios {{ display: grid; grid-template-columns: repeat(3,1fr); gap: 12px; margin: 20px 0; }}
  .scenario {{ background: var(--card2); border: 1px solid var(--border); border-radius: 10px; padding: 16px; text-align: center; }}
  .scenario-icon {{ font-size: 24px; margin-bottom: 8px; }}
  .scenario-name {{ font-size: 13px; font-weight: 600; margin-bottom: 4px; }}
  .scenario-pct {{ font-size: 28px; font-weight: 800; margin: 8px 0; }}
  .scenario-range {{ font-size: 12px; color: var(--muted); }}
  .bull-scenario {{ border-color: #3fb95044; }} .bull-scenario .scenario-pct {{ color: var(--green); }}
  .neutral-scenario {{ border-color: #d2992244; }} .neutral-scenario .scenario-pct {{ color: var(--yellow); }}
  .bear-scenario {{ border-color: #f8514944; }} .bear-scenario .scenario-pct {{ color: var(--red); }}
  .prob-bar-wrap {{ height: 6px; background: #21262d; border-radius: 3px; margin-top: 8px; overflow: hidden; }}
  .prob-bar {{ height: 100%; border-radius: 3px; }}
  .bull-bar {{ background: var(--green); }} .neutral-bar {{ background: var(--yellow); }} .bear-bar {{ background: var(--red); }}
  .rsi-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 12px; margin: 16px 0; }}
  .rsi-item {{ background: var(--card2); border: 1px solid var(--border); border-radius: 8px; padding: 14px; }}
  .rsi-label {{ font-size: 11px; color: var(--muted); font-weight: 600; margin-bottom: 8px; }}
  .rsi-bar-wrap {{ height: 8px; background: linear-gradient(to right,var(--blue) 0%,var(--green) 50%,var(--red) 100%); border-radius: 4px; position: relative; }}
  .rsi-marker {{ width: 14px; height: 14px; background: white; border-radius: 50%; position: absolute; top: -3px; transform: translateX(-50%); border: 2px solid var(--card2); }}
  .rsi-value {{ font-size: 22px; font-weight: 700; margin-top: 8px; }}
  .rsi-desc {{ font-size: 12px; color: var(--muted); }}
  .ema-table {{ width: 100%; border-collapse: collapse; font-size: 13px; margin-top: 12px; }}
  .ema-table th {{ background: #21262d; padding: 8px 12px; text-align: left; font-size: 11px; color: var(--muted); text-transform: uppercase; }}
  .ema-table td {{ padding: 8px 12px; border-bottom: 1px solid #21262d; }}
  .above {{ color: var(--red); }} .below {{ color: var(--green); }}
  .strategy-box {{ background: #1f6feb11; border: 1px solid #1f6feb33; border-radius: 8px; padding: 16px; margin-top: 16px; }}
  .strategy-box h4 {{ font-size: 13px; font-weight: 600; color: var(--blue); margin-bottom: 8px; }}
  .strategy-box p {{ font-size: 13px; color: var(--muted); line-height: 1.7; }}
  .disclaimer {{ background: #9e6a0311; border: 1px solid #9e6a0333; color: var(--yellow); padding: 10px 16px; border-radius: 8px; font-size: 13px; margin-bottom: 20px; }}
  .updated-badge {{ background: #1f6feb22; border: 1px solid #1f6feb55; color: var(--blue); padding: 2px 8px; border-radius: 4px; font-size: 11px; font-weight: 600; }}
  @media (max-width: 768px) {{
    .main {{ padding: 16px; }}
    .charts-grid,.scenarios {{ grid-template-columns: 1fr; }}
    .header {{ padding: 16px; }}
  }}
</style>
</head>
<body>

<div class="header">
  <div class="header-icon">📊</div>
  <div>
    <h1>Stock Portfolio Dashboard</h1>
    <p>Portfolio · AI Trading · P&L · Updated {updated_at}</p>
  </div>
  <div class="badge">🤖 AI Updated</div>
</div>

<div class="tabs">
  <div class="tab {'active' if default_tab == 'portfolio' else ''}" onclick="switchTab('portfolio')">📈 Portfolio</div>
  <div class="tab {'active' if default_tab == 'ai' else ''}" onclick="switchTab('ai')">🤖 AI Trading</div>
</div>

<!-- SHEET 1: PORTFOLIO -->
<div id="tab-portfolio" class="section {'active' if default_tab == 'portfolio' else ''} main">
  <div class="summary-grid">
    <div class="card">
      <div class="card-label">💰 Total Portfolio Value (THB)</div>
      <div class="card-value blue-text">฿{total_val_thb:,.0f}</div>
      <div class="card-sub">≈ ${total_val_usd:,.2f} USD</div>
    </div>
    <div class="card">
      <div class="card-label">📉 Total P&L (USD)</div>
      <div class="card-value {pnl_class_total}">{pnl_sign_total}${abs(total_pnl_usd):,.2f}</div>
      <div class="card-sub {pnl_class_total}">{pnl_sign_total}{total_pnl_pct:.1f}% overall</div>
    </div>
    <div class="card">
      <div class="card-label">💱 Exchange Rate</div>
      <div class="card-value blue-text">{usdthb:.2f}</div>
      <div class="card-sub">USD / THB</div>
    </div>
    <div class="card">
      <div class="card-label">🏢 Holdings</div>
      <div class="card-value blue-text">{len(portfolio)}</div>
      <div class="card-sub">stocks in portfolio</div>
    </div>
    <div class="card">
      <div class="card-label">✅ Gainers</div>
      <div class="card-value positive">{len(gainers)}</div>
      <div class="card-sub positive">{', '.join(r['symbol'] for r in sorted(gainers, key=lambda x: x['pnl_pct'], reverse=True)[:4])}</div>
    </div>
    <div class="card">
      <div class="card-label">❌ Losers</div>
      <div class="card-value negative">{len(losers)}</div>
      <div class="card-sub negative">stocks in portfolio</div>
    </div>
  </div>

  <div class="charts-grid">
    <div class="chart-card">
      <h3>📊 Portfolio Allocation (% by Value)</h3>
      <div class="chart-container"><canvas id="donutChart"></canvas></div>
    </div>
    <div class="chart-card">
      <h3>📉 P&L per Stock (USD)</h3>
      <div class="chart-container"><canvas id="plChart"></canvas></div>
    </div>
  </div>

  <div class="filter-bar">
    <button class="filter-btn active" onclick="filterStocks('all',this)">🔘 All</button>
    <button class="filter-btn" onclick="filterStocks('profit',this)">✅ Gainers</button>
    <button class="filter-btn" onclick="filterStocks('loss',this)">❌ Losers</button>
    <button class="filter-btn" onclick="filterStocks('advice-hold',this)">🔵 Hold</button>
    <button class="filter-btn" onclick="filterStocks('advice-reduce',this)">⚠️ Reduce/Cut</button>
  </div>

  <div class="section-title"><div class="dot"></div> Individual Stock Details</div>
  <div class="table-wrap">
    <table id="stockTable">
      <thead><tr>
        <th>#</th><th>Stock</th><th>Cost/Share ($)</th><th>Current Price ($)</th>
        <th>P&L ($)</th><th>% Change</th><th>Value (USD)</th>
        <th>Value (THB)</th><th>Weight %</th><th>Advice</th><th>Updated</th>
      </tr></thead>
      <tbody id="stockBody">{table_rows_html}</tbody>
    </table>
  </div>

  <div class="section-title"><div class="dot"></div> Portfolio Positions — Value per Stock (THB)</div>
  <div class="chart-card" style="margin-bottom:32px;">
    <div class="chart-container" style="height:320px;"><canvas id="positionChart"></canvas></div>
  </div>
</div>

<!-- SHEET 2: AI TRADING -->
<div id="tab-ai" class="section {'active' if default_tab == 'ai' else ''} main">
  <div class="disclaimer">⚠️ Analysis data {updated_at} — Not investment advice</div>

  <div class="summary-grid" style="margin-bottom:24px;">
    <div class="card">
      <div class="card-label">📊 Week Trades</div>
      <div class="card-value blue-text" style="font-size:20px">{syms_display}</div>
      <div class="card-sub">Hold 3-5 days</div>
    </div>
    <div class="card">
      <div class="card-label">💼 # Positions</div>
      <div class="card-value blue-text">{len(week_trades)}</div>
      <div class="card-sub">Active trades</div>
    </div>
    <div class="card">
      <div class="card-label">💰 Total Budget</div>
      <div class="card-value positive">${budget_total:,.0f}</div>
      <div class="card-sub">Week trade capital</div>
    </div>
    <div class="card">
      <div class="card-label">📥 Total Invested</div>
      <div class="card-value blue-text">${total_invest:,.0f}</div>
      <div class="card-sub">Deployed</div>
    </div>
    <div class="card">
      <div class="card-label">💵 Cash Remaining</div>
      <div class="card-value {'positive' if budget_total - total_invest >= 0 else 'negative'}">${budget_total - total_invest:,.0f}</div>
      <div class="card-sub">Available</div>
    </div>
    <div class="card">
      <div class="card-label">📅 Updated</div>
      <div class="card-value" style="font-size:16px">{updated_at}</div>
      <div class="card-sub">Last refresh</div>
    </div>
  </div>

{trade_cards_html}
</div>

<script>
// ─── Filter ───
function filterStocks(filter, btn) {{
  document.querySelectorAll('.filter-btn').forEach(b => b.classList.remove('active'));
  btn.classList.add('active');
  document.querySelectorAll('#stockBody tr').forEach(row => {{
    if (filter === 'all') {{ row.style.display = ''; return; }}
    const adv = row.dataset.advice || '';
    const pnl = parseFloat(row.cells[4]?.textContent?.replace(/[+$]/,'') || 0);
    if (filter === 'profit') row.style.display = pnl >= 0 ? '' : 'none';
    else if (filter === 'loss') row.style.display = pnl < 0 ? '' : 'none';
    else row.style.display = adv.includes(filter) ? '' : 'none';
  }});
}}

// ─── Charts ───
const donutCtx = document.getElementById('donutChart').getContext('2d');
new Chart(donutCtx, {{
  type: 'doughnut',
  data: {{
    labels: {donut_labels},
    datasets: [{{ data: {donut_data}, backgroundColor: {donut_colors}, borderWidth: 2, borderColor: '#0d1117' }}]
  }},
  options: {{
    responsive: true, maintainAspectRatio: false,
    plugins: {{
      legend: {{ position: 'right', labels: {{ color: '#8b949e', font: {{ size: 11 }} }} }},
      tooltip: {{ callbacks: {{ label: ctx => `${{ctx.label}}: $${{ctx.raw.toFixed(2)}}` }} }}
    }}
  }}
}});

const plCtx = document.getElementById('plChart').getContext('2d');
new Chart(plCtx, {{
  type: 'bar',
  data: {{
    labels: {pl_labels},
    datasets: [{{ label: 'P&L (USD)', data: {pl_data}, backgroundColor: {pl_colors}, borderColor: {pl_borders}, borderWidth: 1, borderRadius: 4 }}]
  }},
  options: {{
    responsive: true, maintainAspectRatio: false,
    plugins: {{ legend: {{ display: false }}, tooltip: {{ callbacks: {{ label: ctx => `$${{ctx.raw.toFixed(2)}}` }} }} }},
    scales: {{
      x: {{ ticks: {{ color: '#8b949e', font: {{ size: 10 }} }}, grid: {{ color: '#21262d' }} }},
      y: {{ ticks: {{ color: '#8b949e', callback: v => `$${{v}}` }}, grid: {{ color: '#21262d' }} }}
    }}
  }}
}});

const posCtx = document.getElementById('positionChart').getContext('2d');
new Chart(posCtx, {{
  type: 'bar',
  data: {{
    labels: {pos_labels},
    datasets: [{{ label: 'Value (THB)', data: {pos_data}, backgroundColor: '#58a6ff44', borderColor: '#58a6ff', borderWidth: 1, borderRadius: 4 }}]
  }},
  options: {{
    responsive: true, maintainAspectRatio: false,
    plugins: {{ legend: {{ display: false }}, tooltip: {{ callbacks: {{ label: ctx => `฿${{ctx.raw.toLocaleString()}}` }} }} }},
    scales: {{
      x: {{ ticks: {{ color: '#8b949e' }}, grid: {{ color: '#21262d' }} }},
      y: {{ ticks: {{ color: '#8b949e', callback: v => `฿${{(v/1000).toFixed(0)}}K` }}, grid: {{ color: '#21262d' }} }}
    }}
  }}
}});

// ─── Scenario Chart (per-trade inline bars — no canvas needed) ───
function initScenarioChart() {{ /* no-op: scenarios rendered inline per trade card */ }}

// ─── Tab Switch ───
function switchTab(tab) {{
  document.querySelectorAll('.tab').forEach((t,i) => {{
    t.classList.toggle('active', (i===0&&tab==='portfolio')||(i===1&&tab==='ai'));
  }});
  document.getElementById('tab-portfolio').classList.toggle('active', tab==='portfolio');
  document.getElementById('tab-ai').classList.toggle('active', tab==='ai');
  if (tab === 'ai') initScenarioChart();
}}
</script>
</body>
</html>"""
    return html


def read_week_trades(ws2, stocks_data: dict) -> list:
    """Read week trade stocks from AI Trading sheet rows 4-9"""
    import re
    trades = []
    budget_total = ws2.cell(row=2, column=3).value or 0
    for row_idx in range(4, 10):
        sym = ws2.cell(row=row_idx, column=2).value
        if not sym or not isinstance(sym, str) or not sym.strip():
            continue
        sym = sym.strip()
        if not re.match(r'^[A-Z]{1,6}$', sym):
            continue
        company  = ws2.cell(row=row_idx, column=3).value or ""
        pe       = ws2.cell(row=row_idx, column=4).value or "N/A"
        entry    = ws2.cell(row=row_idx, column=5).value or 0
        shares   = ws2.cell(row=row_idx, column=6).value or 0
        invest   = ws2.cell(row=row_idx, column=7).value or 0
        sl       = ws2.cell(row=row_idx, column=8).value or 0
        sl_pct   = ws2.cell(row=row_idx, column=9).value or 0
        tp1      = ws2.cell(row=row_idx, column=10).value or 0
        tp2      = ws2.cell(row=row_idx, column=11).value or 0
        tp3      = ws2.cell(row=row_idx, column=12).value or 0
        rr       = ws2.cell(row=row_idx, column=13).value or "N/A"
        strategy = ws2.cell(row=row_idx, column=15).value or ""
        price    = stocks_data.get(sym, {}).get("price", float(entry) if entry else 0)
        rsi1d    = stocks_data.get(sym, {}).get("rsi_1d", 50)
        rsi1h    = stocks_data.get(sym, {}).get("rsi_1h", 50)
        trades.append({
            "sym": sym, "company": str(company), "pe": str(pe),
            "entry": float(entry) if entry else 0,
            "shares": float(shares) if shares else 0,
            "invest": float(invest) if invest else 0,
            "sl": float(sl) if sl else 0,
            "sl_pct": float(sl_pct) if sl_pct else 0,
            "tp1": float(tp1) if tp1 else 0,
            "tp2": float(tp2) if tp2 else 0,
            "tp3": float(tp3) if tp3 else 0,
            "rr": str(rr), "strategy": str(strategy),
            "price": float(price),
            "rsi1d": float(rsi1d), "rsi1h": float(rsi1h),
            "budget_total": float(budget_total),
        })
    return trades


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel",       required=True)
    parser.add_argument("--output",      required=True)
    parser.add_argument("--data",        required=True)
    parser.add_argument("--default-tab", choices=["portfolio", "ai"], default="portfolio",
                        help="Which tab is active when the dashboard first opens "
                             "(portfolio = Sheet 1, ai = Sheet 2 AI Trading)")
    args = parser.parse_args()

    print(f"🌐 Building Dashboard: {args.output}")

    with open(args.data, encoding="utf-8") as f:
        data = json.load(f)

    stocks_data = data.get("stocks", {})
    usdthb      = data.get("usdthb", 33.0)
    updated_at  = data.get("updated_at", "")

    # Read portfolio + week trades from Excel
    wb = openpyxl.load_workbook(args.excel)
    ws = wb["Portfolio"]
    ws2 = wb["AI Trading"]
    portfolio   = read_portfolio(ws, stocks_data, usdthb)
    week_trades = read_week_trades(ws2, stocks_data)
    data["week_trades"] = week_trades

    print(f"   📋 Stocks in portfolio: {len(portfolio)}")
    print(f"   📊 Week trades: {[t['sym'] for t in week_trades]}")

    html = build_html(portfolio, stocks_data, data, usdthb, updated_at,
                      default_tab=args.default_tab)

    with open(args.output, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"   ✅ Dashboard saved → {args.output}")


if __name__ == "__main__":
    main()
