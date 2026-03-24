#!/usr/bin/env python3
"""
AI Trader — Claude-Driven Mode
Uses data gathered by Claude via WebSearch to update Excel + Dashboard.

Accepts stock data as JSON input instead of fetching from yfinance.
Suitable for running in Cowork environment without network access.

Usage:
  python3 claude_trader.py --excel <path> --prices <prices_json> [--output <html_path>]

prices_json format:
  {
    "NVDA": {"price": 175.5, "rsi_1d": 55.3, "rsi_1h": 42.1,
             "ema200_1d": 155.0, "ema50_1d": 172.0,
             "ema20_1h": 177.0, "ema50_1h": 175.0,
             "ema100_1h": 173.0, "ema200_1h": 165.0},
    ...
  }
"""

import argparse
import json
import sys
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment
import warnings
warnings.filterwarnings('ignore')

TODAY_SHORT = datetime.now().strftime("%-d %b")
TODAY_FULL = datetime.now().strftime("%d %b %Y")


def rsi_label(rsi: float) -> str:
    if rsi >= 70: return "Overbought ⚠️"
    if rsi <= 30: return "Oversold ⚡"
    if rsi >= 60: return "Neutral-Bullish ✅"
    if rsi <= 40: return "Near Oversold ✅"
    return "Neutral ✅"


def generate_ai_advice(sym, d: dict, cost_basis: float) -> str:
    price   = d.get("price", 0)
    rsi_1d  = d.get("rsi_1d", 50)
    rsi_1h  = d.get("rsi_1h", 50)
    ema50d  = d.get("ema50_1d", price)
    ema200d = d.get("ema200_1d", price)

    pnl_pct = ((price - cost_basis) / cost_basis * 100) if cost_basis > 0 else 0

    if pnl_pct < -40:
        action = "⚠️ Cut Loss / Reduce"
    elif pnl_pct < -15:
        action = "⚠️ Reduce Position"
    elif rsi_1d > 75:
        action = "[Consider selling partial]"
    elif rsi_1d < 35 and pnl_pct > -10:
        action = "[Hold / Add]"
    elif pnl_pct > 30:
        action = "[Hold / Take Profit partial]"
    else:
        action = "[Hold]"

    above_ema200 = price > ema200d
    ema_trend = "Above EMA200(1D) ✅" if above_ema200 else "Below EMA200(1D) ⚠️"

    lines = [
        f"{action}",
        f"RSI(1D)={rsi_1d} {rsi_label(rsi_1d)} | RSI(1H)={rsi_1h} {rsi_label(rsi_1h)}",
        f"EMA50(1D)=${ema50d:.2f} | EMA200(1D)=${ema200d:.2f}",
        f"Current price {ema_trend} | P&L {pnl_pct:+.1f}%",
        f"📅 Updated {TODAY_SHORT} — auto data",
    ]
    return "\n".join(lines)


def generate_trade_analysis(sym, d: dict) -> str:
    price   = d.get("price", 0)
    rsi_1d  = d.get("rsi_1d", 50)
    rsi_1h  = d.get("rsi_1h", 50)
    ema20h  = d.get("ema20_1h", price)
    ema50h  = d.get("ema50_1h", price)
    ema100h = d.get("ema100_1h", price)
    ema200h = d.get("ema200_1h", price)

    if rsi_1d < 40:
        bull_pct, neutral_pct, bear_pct = 50, 30, 20
        bull_range = f"${price*1.05:.0f}-${price*1.12:.0f}"
        bear_range = f"below ${price*0.95:.0f}"
    elif rsi_1d > 65:
        bull_pct, neutral_pct, bear_pct = 30, 35, 35
        bull_range = f"${price*1.03:.0f}-${price*1.08:.0f}"
        bear_range = f"below ${price*0.95:.0f}"
    else:
        bull_pct, neutral_pct, bear_pct = 45, 35, 20
        bull_range = f"${price*1.05:.0f}-${price*1.12:.0f}"
        bear_range = f"below ${price*0.95:.0f}"

    entry = round(price, 2)
    tp1   = round(price * 1.05, 2)
    tp2   = round(price * 1.08, 2)
    tp3   = round(price * 1.12, 2)
    sl    = round(price * 0.95, 2)

    return "\n".join([
        f"═══ {sym} Analysis · {TODAY_SHORT} ═══",
        "",
        "📈 RSI",
        f"  • 1D RSI = {rsi_1d} → {rsi_label(rsi_1d)}",
        f"  • 1H RSI = {rsi_1h} → {rsi_label(rsi_1h)}",
        "",
        "📉 EMA Levels (1H)",
        f"  • EMA20  = ${ema20h:.2f}",
        f"  • EMA50  = ${ema50h:.2f}",
        f"  • EMA100 = ${ema100h:.2f}  ← Entry zone",
        f"  • EMA200 = ${ema200h:.2f}  ← Stop Loss zone",
        "",
        "🎯 Scenarios",
        f"  • 🐂 Bullish  {bull_pct}%  → {bull_range}",
        f"  • 📊 Neutral  {neutral_pct}%  → sideways",
        f"  • 🐻 Bearish  {bear_pct}%  → {bear_range}",
        "",
        "💰 Trade Levels",
        f"  • Entry  : ${entry}",
        f"  • TP1    : ${tp1}  (+5%)  Sell 40%",
        f"  • TP2    : ${tp2}  (+8%)  Sell 40%",
        f"  • TP3    : ${tp3}  (+12%) Sell 20%",
        f"  • SL     : ${sl}  (-5%)  Cut all",
        "",
        "⚠️ Not investment advice",
    ])


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel",   required=True, help="Path to Excel file")
    parser.add_argument("--prices",  required=True, help="JSON string or file path with stock prices+indicators")
    parser.add_argument("--usdthb",  type=float, default=0, help="USD/THB rate (0 = keep existing)")
    parser.add_argument("--week-trade", default="", help="Current week trade symbol")
    parser.add_argument("--output",      default="", help="HTML output path (empty = same as Excel folder)")
    parser.add_argument("--default-tab", choices=["portfolio", "ai"], default="portfolio",
                        help="Which dashboard tab opens first (portfolio or ai)")
    args = parser.parse_args()

    # Parse prices
    if args.prices.startswith("{"):
        prices = json.loads(args.prices)
    else:
        with open(args.prices) as f:
            prices = json.load(f)

    print(f"🤖 Claude AI Trader — {TODAY_FULL}")
    print(f"   📥 Stocks received: {len(prices)}")

    # Load Excel
    wb = openpyxl.load_workbook(args.excel)
    ws_port  = wb["Portfolio"]
    ws_trade = wb["AI Trading"]

    # Update USD/THB — always fetch fresh (never reuse old value)
    usdthb = args.usdthb
    if usdthb <= 0:
        print("   💱 Fetching USD/THB from Yahoo Finance...", end=" ", flush=True)
        try:
            import urllib.request
            for sym in ("USDTHB=X", "THB=X"):
                url = (f"https://query1.finance.yahoo.com/v8/finance/chart/{sym}"
                       f"?interval=1d&range=5d")
                req = urllib.request.Request(
                    url, headers={"User-Agent": "Mozilla/5.0"})
                resp = urllib.request.urlopen(req, timeout=10)
                data = json.loads(resp.read())
                closes = data["chart"]["result"][0]["indicators"]["quote"][0]["close"]
                val = float([c for c in closes if c][-1])
                if 25 < val < 50:
                    usdthb = round(val, 2)
                    break
        except Exception:
            pass
        if usdthb <= 0:
            usdthb = 33.0   # fallback if fetch fails
        print(usdthb)
    ws_port.cell(row=2, column=7).value = usdthb
    print(f"   💱 USD/THB = {usdthb} (updated)")

    # Update portfolio sheet
    import re as _re
    VALID_TICKER = _re.compile(r'^[A-Z]{1,6}$')
    updated = 0
    for row_idx in range(5, 70):
        sym_val = ws_port.cell(row=row_idx, column=2).value
        # Stop at first empty cell — marks end of stock table
        if sym_val is None:
            break
        if not isinstance(sym_val, str):
            continue
        sym = sym_val.strip()
        # Skip non-ticker rows (labels, formulas, dashes)
        if not VALID_TICKER.match(sym):
            continue
        if sym in prices:
            d = prices[sym]
            price = d.get("price", 0)
            cost  = ws_port.cell(row=row_idx, column=3).value
            cost_f = float(cost) if isinstance(cost, (int, float)) else 0

            ws_port.cell(row=row_idx, column=4).value = price
            ws_port.cell(row=row_idx, column=9).value = generate_ai_advice(sym, d, cost_f)
            ws_port.cell(row=row_idx, column=9).alignment = Alignment(wrap_text=True)
            ws_port.cell(row=row_idx, column=11).value = f"🆕 {TODAY_SHORT}"
            updated += 1

    print(f"   ✅ Updated {updated} stocks in Excel")

    # Update AI Trading sheet
    if args.week_trade and args.week_trade in prices:
        wt = args.week_trade
        d  = prices[wt]
        for row_idx in range(4, 12):
            sym_val = ws_trade.cell(row=row_idx, column=2).value
            if sym_val and str(sym_val).strip() == wt:
                ws_trade.cell(row=row_idx, column=5).value  = round(d["price"], 2)
                ws_trade.cell(row=row_idx, column=23).value = generate_trade_analysis(wt, d)
                ws_trade.cell(row=row_idx, column=23).alignment = Alignment(wrap_text=True)
                break
        print(f"   ✅ AI Trading updated: {wt}")

    wb.save(args.excel)
    print(f"   💾 Excel saved")

    # Build dashboard
    if args.output:
        html_path = args.output
    else:
        import os
        html_path = os.path.join(os.path.dirname(args.excel), "index.html")

    # Prepare full data structure for build_dashboard
    stocks_full = {}
    for sym, d in prices.items():
        stocks_full[sym] = {**d, "cost_basis": 0, "ai_advice": "", "trade_analysis": ""}

    data_json = {
        "updated_at":    TODAY_FULL,
        "updated_at_th": TODAY_SHORT,
        "usdthb":        usdthb,
        "stocks":        stocks_full,
        "week_trade":    args.week_trade or None,
    }

    import tempfile, os
    tmp_data = tempfile.mktemp(suffix=".json")
    with open(tmp_data, "w") as f:
        json.dump(data_json, f, ensure_ascii=False)

    # Import and run build_dashboard
    script_dir = os.path.dirname(os.path.abspath(__file__))
    sys.path.insert(0, script_dir)
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "build_dashboard",
        os.path.join(script_dir, "build_dashboard.py")
    )
    bd = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(bd)

    wb2 = openpyxl.load_workbook(args.excel)
    ws2 = wb2["Portfolio"]
    portfolio = bd.read_portfolio(ws2, stocks_full, data_json["usdthb"])
    html = bd.build_html(portfolio, stocks_full, data_json, data_json["usdthb"], TODAY_FULL,
                         default_tab=args.default_tab)

    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)

    os.unlink(tmp_data)
    print(f"   🌐 Dashboard rebuilt → {html_path}")
    print(f"\n✅ Done! Files ready:")
    print(f"   📊 Excel: {args.excel}")
    print(f"   🌐 HTML:  {html_path}")


if __name__ == "__main__":
    main()
