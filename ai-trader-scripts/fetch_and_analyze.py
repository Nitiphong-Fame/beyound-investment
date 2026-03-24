#!/usr/bin/env python3
"""
AI Trader — Fetch & Analyze (Yahoo Finance Direct)
Fetch real-time stock prices + calculate RSI/EMA directly from Yahoo Finance API.
No yfinance needed — uses only requests + pandas + numpy

Usage:
  python3 fetch_and_analyze.py \
    --excel "/path/to/รายรับ_รายจ่าย_ประจำปี.xlsx"
"""

import argparse
import json
import sys
import time
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

import requests
import pandas as pd
import numpy as np
import openpyxl

TODAY      = datetime.now().strftime("%d %b %Y")
TODAY_SHORT = datetime.now().strftime("%-d %b")

YF_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json",
}


# ─────────────────────────────────────────────
# Yahoo Finance API Helpers
# ─────────────────────────────────────────────

def yf_chart(symbol: str, interval: str, range_: str, retries: int = 3) -> pd.DataFrame | None:
    """
    Fetch OHLCV from Yahoo Finance Chart API
    interval: "1d" | "1h" | "4h"
    range_:   "1y" | "6mo" | "3mo" | "60d" | "5d"
    """
    # Yahoo Finance uses query1 or query2 (alternate if rate-limited)
    for host in ("query1", "query2"):
        url = (
            f"https://{host}.finance.yahoo.com/v8/finance/chart/{symbol}"
            f"?interval={interval}&range={range_}&includeAdjustedClose=true"
        )
        for attempt in range(retries):
            try:
                resp = requests.get(url, headers=YF_HEADERS, timeout=15)
                if resp.status_code == 429:
                    time.sleep(2 ** attempt)
                    continue
                resp.raise_for_status()
                data = resp.json()
                result = data["chart"]["result"]
                if not result:
                    return None
                r       = result[0]
                ts      = r["timestamp"]
                closes  = r["indicators"]["quote"][0]["close"]
                highs   = r["indicators"]["quote"][0].get("high", closes)
                lows    = r["indicators"]["quote"][0].get("low", closes)
                volumes = r["indicators"]["quote"][0].get("volume", [0]*len(ts))

                df = pd.DataFrame({
                    "timestamp": pd.to_datetime(ts, unit="s"),
                    "close":  [float(c) if c is not None else None for c in closes],
                    "high":   [float(h) if h is not None else None for h in highs],
                    "low":    [float(l) if l is not None else None for l in lows],
                    "volume": [float(v) if v is not None else 0   for v in volumes],
                }).dropna(subset=["close"])

                if len(df) < 5:
                    return None
                return df

            except requests.exceptions.RequestException as e:
                if attempt == retries - 1:
                    raise
                time.sleep(1)
    return None


# ─────────────────────────────────────────────
# Yahoo Finance News
# ─────────────────────────────────────────────

def fetch_news(symbol: str, max_items: int = 3) -> list[dict]:
    """
    Fetch latest news from Yahoo Finance Search API.
    Returns list of {"title": str, "date": str} or [] if no news.
    """
    url = (
        f"https://query1.finance.yahoo.com/v1/finance/search"
        f"?q={symbol}&newsCount={max_items}&enableCb=false&enableNavLinks=false"
    )
    try:
        resp = requests.get(url, headers=YF_HEADERS, timeout=10)
        resp.raise_for_status()
        data = resp.json()
        news_items = data.get("news", [])
        results = []
        for item in news_items[:max_items]:
            title = item.get("title", "").strip()
            if not title:
                continue
            # Convert Unix timestamp to readable date
            ts = item.get("providerPublishTime", 0)
            if ts:
                dt = datetime.fromtimestamp(ts)
                date_str = dt.strftime("%-d %b")
            else:
                date_str = TODAY_SHORT
            results.append({"title": title, "date": date_str})
        return results
    except Exception:
        return []


# ─────────────────────────────────────────────
# Technical Indicators
# ─────────────────────────────────────────────

def calc_rsi(closes: pd.Series, period: int = 14) -> float:
    delta    = closes.diff()
    gain     = delta.clip(lower=0)
    loss     = -delta.clip(upper=0)
    avg_gain = gain.ewm(com=period - 1, min_periods=period).mean()
    avg_loss = loss.ewm(com=period - 1, min_periods=period).mean()
    rs       = avg_gain / avg_loss
    rsi      = 100 - (100 / (1 + rs))
    return round(float(rsi.iloc[-1]), 2)


def calc_ema(closes: pd.Series, period: int) -> float:
    return round(float(closes.ewm(span=period, adjust=False).mean().iloc[-1]), 4)


def rsi_label(rsi: float) -> str:
    if rsi >= 70: return "Overbought ⚠️"
    if rsi <= 30: return "Oversold ⚡ (bounce likely)"
    if rsi >= 60: return "Neutral-Bullish ✅"
    if rsi <= 40: return "Near Oversold ✅"
    return "Neutral ✅"


# ─────────────────────────────────────────────
# Per-Symbol Fetch
# ─────────────────────────────────────────────

def fetch_ticker(symbol: str, is_gold: bool = False) -> dict | None:
    """Fetch data for 1 stock — price + RSI + EMA directly from Yahoo Finance"""
    yf_sym = "GC=F" if is_gold else symbol

    try:
        # Daily (1 year for EMA200)
        df_1d = yf_chart(yf_sym, "1d", "1y")
        if df_1d is None or len(df_1d) < 20:
            print(f"  ⚠️  {symbol}: insufficient daily data", file=sys.stderr)
            return None

        price = round(float(df_1d["close"].iloc[-1]), 4)

        rsi_1d   = calc_rsi(df_1d["close"])
        ema20_1d  = calc_ema(df_1d["close"], 20)
        ema50_1d  = calc_ema(df_1d["close"], 50)
        ema100_1d = calc_ema(df_1d["close"], 100) if len(df_1d) >= 100 else ema50_1d
        ema200_1d = calc_ema(df_1d["close"], 200) if len(df_1d) >= 200 else ema100_1d

        high_52w = round(float(df_1d["high"].tail(252).max()), 4)
        low_52w  = round(float(df_1d["low"].tail(252).min()),  4)
        avg_vol  = float(df_1d["volume"].tail(20).mean())
        last_vol = float(df_1d["volume"].iloc[-1])
        vol_ratio = round(last_vol / avg_vol, 2) if avg_vol > 0 else 1.0

        # Hourly (60 days for intraday EMA)
        try:
            df_1h = yf_chart(yf_sym, "1h", "60d")
            if df_1h is None or len(df_1h) < 20:
                raise ValueError("not enough 1h data")
            rsi_1h    = calc_rsi(df_1h["close"])
            ema20_1h  = calc_ema(df_1h["close"], 20)
            ema50_1h  = calc_ema(df_1h["close"], 50)
            ema100_1h = calc_ema(df_1h["close"], 100) if len(df_1h) >= 100 else ema50_1h
            ema200_1h = calc_ema(df_1h["close"], 200) if len(df_1h) >= 200 else ema100_1h
        except Exception:
            # Fallback: use 1D values
            rsi_1h = rsi_1d
            ema20_1h  = ema20_1d
            ema50_1h  = ema50_1d
            ema100_1h = ema100_1d
            ema200_1h = ema200_1d

        return {
            "symbol":   symbol,
            "price":    price,
            "rsi_1d":   rsi_1d,
            "rsi_1h":   rsi_1h,
            "ema20_1d":  ema20_1d,
            "ema50_1d":  ema50_1d,
            "ema100_1d": ema100_1d,
            "ema200_1d": ema200_1d,
            "ema20_1h":  ema20_1h,
            "ema50_1h":  ema50_1h,
            "ema100_1h": ema100_1h,
            "ema200_1h": ema200_1h,
            "vol_ratio": vol_ratio,
            "high_52w":  high_52w,
            "low_52w":   low_52w,
        }

    except Exception as e:
        print(f"  ⚠️  {symbol}: Yahoo Finance failed ({e}) — trying Google Finance...",
              file=sys.stderr)
        # Google Finance fallback: price only, RSI/EMA use defaults
        gf_price = google_finance_price(symbol)
        if gf_price:
            print(f"  ✅  {symbol}: Google Finance price = ${gf_price}", file=sys.stderr)
            return {
                "symbol":    symbol,
                "price":     gf_price,
                "rsi_1d":    50.0,
                "rsi_1h":    50.0,
                "ema20_1d":  round(gf_price * 1.02, 4),
                "ema50_1d":  round(gf_price * 1.01, 4),
                "ema100_1d": round(gf_price * 0.99, 4),
                "ema200_1d": round(gf_price * 0.97, 4),
                "ema20_1h":  round(gf_price * 1.01, 4),
                "ema50_1h":  round(gf_price * 1.005, 4),
                "ema100_1h": round(gf_price * 0.995, 4),
                "ema200_1h": round(gf_price * 0.98, 4),
                "vol_ratio": 1.0,
                "high_52w":  round(gf_price * 1.3, 4),
                "low_52w":   round(gf_price * 0.7, 4),
                "_source":   "google_finance",
            }
        print(f"  ❌  {symbol}: both Yahoo and Google Finance failed", file=sys.stderr)
        return None


def google_finance_price(symbol: str) -> float | None:
    """
    Fetch current price from Google Finance as fallback when Yahoo Finance is blocked.
    Tries NASDAQ first, then NYSE, NYSEARCA. Returns price or None.
    Only provides the spot price — no RSI/EMA (use defaults for those).
    """
    import re as _re
    exchanges = ["NASDAQ", "NYSE", "NYSEARCA", "BATS", "CURRENCY"]
    url_sym = "USDTHB" if symbol in ("USDTHB=X", "THB=X") else symbol

    for exchange in exchanges:
        url = f"https://www.google.com/finance/quote/{url_sym}:{exchange}"
        try:
            resp = requests.get(
                url,
                headers={
                    "User-Agent": (
                        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) "
                        "Chrome/122.0.0.0 Safari/537.36"
                    ),
                    "Accept-Language": "en-US,en;q=0.9",
                },
                timeout=12,
                allow_redirects=True,
            )
            if resp.status_code == 404:
                continue
            if resp.status_code != 200:
                continue

            # Google Finance embeds price in several places; try structured patterns first
            text = resp.text

            # Pattern 1: JSON-like data embedded in <c-wiz> or data attributes
            m = _re.search(r'"price":\s*"?([\d,]+\.?\d*)"?', text)
            if m:
                val = float(m.group(1).replace(",", ""))
                if val > 0:
                    return round(val, 4)

            # Pattern 2: Large price div (class names vary but price follows $ sign)
            m = _re.search(r'\$\s*([\d,]+\.\d{2})\b', text)
            if m:
                val = float(m.group(1).replace(",", ""))
                if val > 0:
                    return round(val, 4)

            # Pattern 3: data-last-price attribute
            m = _re.search(r'data-last-price="([\d.]+)"', text)
            if m:
                val = float(m.group(1))
                if val > 0:
                    return round(val, 4)

        except Exception:
            continue

    return None


def fetch_usdthb() -> float:
    """Fetch USD/THB — tries Yahoo Finance first, then Google Finance"""
    # Try Yahoo Finance
    for sym in ("USDTHB=X", "THB=X"):
        try:
            df = yf_chart(sym, "1d", "5d")
            if df is not None and len(df) > 0:
                val = float(df["close"].iloc[-1])
                if 25 < val < 50:   # sanity check
                    return round(val, 2)
        except Exception:
            pass

    # Fallback: Google Finance
    val = google_finance_price("USDTHB=X")
    if val and 25 < val < 50:
        return round(val, 2)

    return 33.0  # last resort default


# ─────────────────────────────────────────────
# AI Advice Generators
# ─────────────────────────────────────────────

def generate_ai_advice(d: dict, cost_basis: float) -> str:
    sym      = d["symbol"]
    price    = d["price"]
    rsi_1d   = d["rsi_1d"]
    rsi_1h   = d["rsi_1h"]
    pnl_pct  = ((price - cost_basis) / cost_basis * 100) if cost_basis > 0 else 0
    news     = d.get("news", [])   # list of {"title": str, "date": str}

    if pnl_pct < -40:
        action = "⚠️ Cut Loss / Reduce"
    elif pnl_pct < -15:
        action = "⚠️ Reduce Position"
    elif rsi_1d > 75:
        action = "[Consider selling partial]"
    elif rsi_1d < 35 and pnl_pct > -10:
        action = "[Hold / Add]"
    elif pnl_pct > 30:
        action = "[Hold / Take Profit partial]"
    else:
        action = "[Hold]"

    above_ema200 = price > d["ema200_1d"]
    ema_trend    = "Above EMA200(1D) ✅" if above_ema200 else "Below EMA200(1D) ⚠️"

    lines = [
        f"{action}",
        f"RSI(1D)={rsi_1d} {rsi_label(rsi_1d)} | RSI(1H)={rsi_1h} {rsi_label(rsi_1h)}",
        f"EMA50(1D)=${d['ema50_1d']:.2f} | EMA200(1D)=${d['ema200_1d']:.2f}",
        f"Current price {ema_trend} | P&L {pnl_pct:+.1f}%",
    ]

    # Add latest news only when available
    if news:
        lines.append(f"📰 Latest news ({TODAY_SHORT}):")
        for item in news:
            lines.append(f"• [{item['date']}] {item['title']}")

    lines.append(f"📅 Updated {TODAY_SHORT} — Yahoo Finance")
    return "\n".join(lines)


def generate_trade_analysis(d: dict) -> str:
    sym    = d["symbol"]
    price  = d["price"]
    rsi_1d = d["rsi_1d"]
    rsi_1h = d["rsi_1h"]

    if rsi_1d < 40:
        bull_pct, neutral_pct, bear_pct = 50, 30, 20
    elif rsi_1d > 65:
        bull_pct, neutral_pct, bear_pct = 30, 35, 35
    else:
        bull_pct, neutral_pct, bear_pct = 45, 35, 20

    return "\n".join([
        f"═══ {sym} Analysis · {TODAY_SHORT} ═══",
        "",
        "📈 RSI",
        f"  • 1D RSI = {rsi_1d} → {rsi_label(rsi_1d)}",
        f"  • 1H RSI = {rsi_1h} → {rsi_label(rsi_1h)}",
        "",
        "📉 EMA Levels (1H)",
        f"  • EMA20  = ${d['ema20_1h']:.2f}",
        f"  • EMA50  = ${d['ema50_1h']:.2f}",
        f"  • EMA100 = ${d['ema100_1h']:.2f}  ← Entry zone",
        f"  • EMA200 = ${d['ema200_1h']:.2f}  ← Stop Loss zone",
        "",
        "🎯 Scenarios",
        f"  • 🐂 Bullish  {bull_pct}%  → ${price*1.05:.0f}–${price*1.12:.0f}",
        f"  • 📊 Neutral  {neutral_pct}%  → sideways",
        f"  • 🐻 Bearish  {bear_pct}%  → below ${price*0.95:.0f}",
        "",
        "💰 Trade Levels",
        f"  • Entry  : ${round(price, 2)}",
        f"  • TP1    : ${round(price*1.05, 2)}  (+5%)  Sell 40%",
        f"  • TP2    : ${round(price*1.08, 2)}  (+8%)  Sell 40%",
        f"  • TP3    : ${round(price*1.12, 2)}  (+12%) Sell 20%",
        f"  • SL     : ${round(price*0.95, 2)}  (-5%)  Cut all",
        "",
        "⚠️ Not investment advice",
    ])


# ─────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel",  required=True, help="Path to Excel file")
    parser.add_argument("--output", default="/tmp/ai_trader_data.json")
    parser.add_argument("--delay",  type=float, default=0.5,
                        help="Delay between requests (seconds) to avoid rate-limit")
    args = parser.parse_args()

    print(f"🤖 AI Trader — {TODAY}  (Yahoo Finance Direct)")
    print("=" * 60)

    # Read portfolio from Excel
    wb       = openpyxl.load_workbook(args.excel)
    ws_port  = wb["Portfolio"]
    ws_trade = wb["AI Trading"]

    portfolio = []
    for row in ws_port.iter_rows(min_row=5, max_row=60, values_only=True):
        sym  = row[1]
        cost = row[2]
        val  = row[6]
        if sym and isinstance(sym, str) and sym.strip():
            portfolio.append({
                "symbol":     sym.strip(),
                "cost_basis": float(cost) if isinstance(cost, (int, float)) else 0,
                "value_usd":  float(val)  if isinstance(val,  (int, float)) else 0,
            })

    # Week trade from AI Trading sheet
    week_trade_sym = None
    for row in ws_trade.iter_rows(min_row=4, max_row=10, values_only=True):
        if row[1] and isinstance(row[1], str) and row[1].strip() not in ("Symbol", "หุ้น"):
            week_trade_sym = row[1].strip()
            break

    print(f"📋 Portfolio: {len(portfolio)} stocks  |  Week Trade: {week_trade_sym}")

    # Fetch USD/THB
    print("\n💱 USD/THB...", end=" ", flush=True)
    usdthb = fetch_usdthb()
    print(usdthb)
    time.sleep(args.delay)

    # Fetch all stock prices
    results = {}
    print("\n📥 Fetching prices + calculating RSI/EMA:")
    all_symbols = list({p["symbol"] for p in portfolio})
    if week_trade_sym and week_trade_sym not in all_symbols:
        all_symbols.append(week_trade_sym)

    for item in portfolio:
        sym     = item["symbol"]
        is_gold = (sym == "GOLD")
        print(f"   {sym:<6}", end=" ", flush=True)

        data = fetch_ticker(sym, is_gold=is_gold)
        if data:
            data["cost_basis"] = item["cost_basis"]

            # Fetch latest news (GOLD has no news in Yahoo Finance search)
            news = [] if is_gold else fetch_news(sym)
            data["news"] = news

            data["ai_advice"]      = generate_ai_advice(data, item["cost_basis"])
            data["trade_analysis"] = generate_trade_analysis(data)
            results[sym] = data

            pnl = ((data["price"] - item["cost_basis"]) / item["cost_basis"] * 100) if item["cost_basis"] > 0 else 0
            sign = "+" if pnl >= 0 else ""
            news_tag = f"  📰{len(news)} news" if news else ""
            print(f"${data['price']:.2f}  RSI1D={data['rsi_1d']:5.1f}  EMA200=${data['ema200_1d']:.2f}  P&L={sign}{pnl:.1f}%{news_tag}")
        else:
            print("skipped (fetch error)")

        time.sleep(args.delay)

    # Week trade if not in portfolio
    if week_trade_sym and week_trade_sym not in results:
        print(f"   {week_trade_sym:<6}", end=" ", flush=True)
        data = fetch_ticker(week_trade_sym)
        if data:
            data["cost_basis"]     = 0
            data["news"]           = fetch_news(week_trade_sym)
            data["ai_advice"]      = generate_ai_advice(data, 0)
            data["trade_analysis"] = generate_trade_analysis(data)
            results[week_trade_sym] = data
            print(f"${data['price']:.2f}  RSI1D={data['rsi_1d']:5.1f}")
        else:
            print("skipped")

    # Summary
    print("\n" + "=" * 60)
    ranked = sorted(
        [(s, (d["price"]-d["cost_basis"])/d["cost_basis"]*100)
         for s, d in results.items() if d["cost_basis"] > 0],
        key=lambda x: x[1], reverse=True
    )
    print("🏆 Top Gainers:", "  ".join(f"{s} {p:+.1f}%" for s, p in ranked[:3]))
    print("📉 Top Losers: ", "  ".join(f"{s} {p:+.1f}%" for s, p in ranked[-3:]))
    print(f"\n✅ Fetched: {len(results)}/{len(portfolio)} stocks  |  USD/THB: {usdthb}")

    # Save JSON
    output = {
        "updated_at":    TODAY,
        "updated_at_th": TODAY_SHORT,
        "usdthb":        usdthb,
        "stocks":        results,
        "week_trade":    week_trade_sym,
    }
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f"💾 Saved to {args.output}")


if __name__ == "__main__":
    main()
