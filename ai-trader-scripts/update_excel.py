#!/usr/bin/env python3
"""
AI Trader — Update Excel
Update Excel file with latest prices + AI Advice from fetch_and_analyze
"""

import argparse
import json
import sys
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import warnings
warnings.filterwarnings('ignore')

TODAY_SHORT = datetime.now().strftime("%-d %b")


def update_portfolio_sheet(ws, stocks: dict, usdthb: float, updated_at: str):
    """Update portfolio sheet (พอร์ตหุ้นรายตัว)"""
    updated = 0
    skipped = 0

    # Update USD/THB (row 2, col G = column 7)
    ws.cell(row=2, column=7).value = usdthb

    import re as _re
    # Get set of merged cell ranges (master cells only are writable)
    merged_ranges = ws.merged_cells.ranges

    def is_merged_non_master(row, col):
        """Return True if cell is part of a merged range but NOT the master cell."""
        from openpyxl.utils import get_column_letter
        cell = ws.cell(row=row, column=col)
        return hasattr(cell, 'value') is False or type(cell).__name__ == 'MergedCell'

    def safe_set(row, col, value, wrap=False):
        try:
            c = ws.cell(row=row, column=col)
            c.value = value
            if wrap:
                c.alignment = Alignment(wrap_text=True)
        except AttributeError:
            pass  # merged cell — skip silently

    for row_idx in range(5, 70):
        sym_cell = ws.cell(row=row_idx, column=2).value
        if not sym_cell or not isinstance(sym_cell, str):
            continue
        sym = sym_cell.strip()
        # Skip header/formula/special rows
        if sym in ("หุ้น", "รวม", "ชื่อ", "Stock", "20 Stocks") or sym.startswith("=") or sym == "—":
            continue

        if sym in stocks:
            d = stocks[sym]
            price = d["price"]

            # Col D = current price (column 4)
            safe_set(row_idx, 4, price)

            # Col I = AI Advice (column 9)
            safe_set(row_idx, 9, d.get("ai_advice", ""), wrap=True)

            # Col K = updated date (column 11)
            safe_set(row_idx, 11, f"🆕 {TODAY_SHORT}")

            updated += 1
        else:
            skipped += 1

    print(f"   ✅ Portfolio: updated {updated} stocks, skipped {skipped}")


def update_portfolio_summary(ws, stocks: dict, usdthb: float, updated_at: str):
    """
    Update summary analysis section (rows 36-47) in portfolio sheet.
    Does not delete template — leaves row blank if no stocks in that category.
    """

    # Calculate P&L per stock and group
    cut_loss  = []   # P&L < -20%
    watch     = []   # P&L -10% to -20%
    profit    = []   # P&L > 0%
    neutral   = []   # P&L -10% to 0%

    total_val_usd = 0
    total_pnl_usd = 0

    for sym, d in stocks.items():
        price = d.get("price", 0)
        cost  = d.get("cost_basis", 0)
        val   = d.get("value_usd", 0)
        if cost <= 0 or price <= 0:
            continue

        pnl_pct = (price - cost) / cost * 100
        val_usd = float(val) if isinstance(val, (int, float)) and val > 0 else 0

        # Calculate P&L USD from value
        if val_usd > 0:
            pnl_usd = val_usd * (price - cost) / price
            total_val_usd += val_usd
            total_pnl_usd += pnl_usd
        else:
            pnl_usd = 0

        item = {
            "sym": sym, "price": price, "pnl_pct": pnl_pct,
            "pnl_usd": pnl_usd, "val_usd": val_usd,
        }

        if pnl_pct < -20:
            cut_loss.append(item)
        elif pnl_pct < -10:
            watch.append(item)
        elif pnl_pct > 0:
            profit.append(item)
        else:
            neutral.append(item)

    # Sort by P&L
    cut_loss.sort(key=lambda x: x["pnl_pct"])
    watch.sort(key=lambda x: x["pnl_pct"])
    profit.sort(key=lambda x: x["pnl_pct"], reverse=True)
    neutral.sort(key=lambda x: x["pnl_pct"])

    total_cost_usd   = total_val_usd - total_pnl_usd
    total_pnl_pct    = (total_pnl_usd / total_cost_usd * 100) if total_cost_usd > 0 else 0
    total_val_thb    = total_val_usd * usdthb
    gainers_str      = "  ".join(f"{x['sym']}({x['pnl_pct']:+.1f}%)" for x in profit[:4])
    action_items     = []
    if cut_loss:
        action_items.append(f"1) CUT {cut_loss[0]['sym']} urgent ({cut_loss[0]['pnl_pct']:.0f}%)")
    if watch:
        action_items.append(f"2) Watch {watch[0]['sym']} ({watch[0]['pnl_pct']:.0f}%)")
    if profit:
        action_items.append(f"3) {profit[0]['sym']}/{profit[1]['sym'] if len(profit)>1 else ''} Hold/Add")

    def fmt_cut(x):
        return (f"• {x['sym']} ({x['pnl_pct']:+.1f}%)  Price ${x['price']:.2f}"
                f"  |  Loss P&L ${x['pnl_usd']:+.0f}"
                f"  |  See AI Advice for details")

    def fmt_watch(x):
        return (f"• {x['sym']} ({x['pnl_pct']:+.1f}%)  ${x['price']:.2f}"
                f"  |  Monitor closely — see AI Advice")

    def fmt_profit(x):
        return (f"• {x['sym']} ({x['pnl_pct']:+.1f}%)  ${x['price']:.2f}"
                f"  |  Hold")

    # ─── Row 36: Header ───
    ws.cell(row=36, column=1).value = (
        f"🤖  Portfolio Analysis Summary — based on prices {updated_at}"
        f"  |  USD/THB {usdthb}"
    )

    # ─── Row 37: CUT LOSS header ───
    n_cut = len(cut_loss)
    ws.cell(row=37, column=1).value = (
        f"🔴  ✂️  CUT LOSS urgent — {n_cut} stocks" if n_cut > 0
        else "🔴  ✂️  CUT LOSS urgent — 0 stocks (none currently)"
    )

    # ─── Row 38: CUT LOSS list (blank if none) ───
    ws.cell(row=38, column=1).value = (
        "\n".join(fmt_cut(x) for x in cut_loss) if cut_loss else ""
    )
    ws.cell(row=38, column=1).alignment = Alignment(wrap_text=True)

    # ─── Row 39: WATCH header ───
    n_watch = len(watch)
    ws.cell(row=39, column=1).value = (
        f"🟡  👀  Watch closely — {n_watch} stocks  (loss 10-20%)" if n_watch > 0
        else "🟡  👀  Watch closely — 0 stocks"
    )

    # ─── Row 40: WATCH list ───
    ws.cell(row=40, column=1).value = (
        "\n".join(fmt_watch(x) for x in watch) if watch else ""
    )
    ws.cell(row=40, column=1).alignment = Alignment(wrap_text=True)

    # ─── Row 41: PROFIT header ───
    n_profit = len(profit)
    ws.cell(row=41, column=1).value = (
        f"🟢  💪  Strong / Profitable — {n_profit} stocks  amid Market"
    )

    # ─── Row 42: PROFIT list ───
    ws.cell(row=42, column=1).value = (
        "\n".join(fmt_profit(x) for x in profit) if profit else ""
    )
    ws.cell(row=42, column=1).alignment = Alignment(wrap_text=True)

    # ─── Row 43: NEUTRAL header ───
    ws.cell(row=43, column=1).value = "⚖️  Near breakeven / small gain-loss — Hold, no action needed"

    # ─── Row 44-45: NEUTRAL list (split into 2 lines) ───
    neutral_strs = [f"{x['sym']} ({x['pnl_pct']:+.1f}%) ${x['price']:.2f}" for x in neutral]
    mid = (len(neutral_strs) + 1) // 2
    ws.cell(row=44, column=1).value = "  |  ".join(neutral_strs[:mid])
    ws.cell(row=45, column=1).value = "  |  ".join(neutral_strs[mid:]) if len(neutral_strs) > mid else ""

    # ─── Row 47: Overall summary ───
    ws.cell(row=47, column=1).value = (
        f"📌  Summary: Total portfolio ${total_val_usd:,.0f} USD"
        f"  ({total_pnl_pct:+.1f}% vs cost)"
        f"  ≈ {total_val_thb:,.0f} THB (rate {usdthb})"
        f"  |  Gainers {n_profit}: {gainers_str}"
        f"  |  ⚡ Action: {'  '.join(action_items) if action_items else 'Hold entire portfolio'}"
    )
    ws.cell(row=47, column=1).alignment = Alignment(wrap_text=True)

    print(f"   ✅ Portfolio summary: CUT={n_cut}  WATCH={n_watch}  PROFIT={n_profit}  NEUTRAL={len(neutral)}")


def update_ai_trading_sheet(ws, stocks: dict, week_trade: str, updated_at: str):
    """Update AI Trading sheet"""
    # Update header timestamp
    header_cell = ws.cell(row=1, column=1)
    if header_cell.value:
        import re
        header = str(header_cell.value)
        header = re.sub(r'(Updated:).*?(\s*\||\s*$)', f'Updated: {updated_at}\\2', header)
        header_cell.value = header

    # Update entry price for week trade
    for row_idx in range(4, 12):
        sym_cell = ws.cell(row=row_idx, column=2).value
        if not sym_cell or not isinstance(sym_cell, str):
            continue
        sym = sym_cell.strip()

        if sym in stocks:
            d = stocks[sym]
            price = d["price"]

            # Col E = Entry price (column 5) — update only if it's the week trade
            if sym == week_trade:
                ws.cell(row=row_idx, column=5).value = round(price, 2)

            # Col W = Analysis Summary (column 23)
            ws.cell(row=row_idx, column=23).value = d.get("trade_analysis", "")
            ws.cell(row=row_idx, column=23).alignment = Alignment(wrap_text=True)

    print(f"   ✅ AI Trading: updated {week_trade if week_trade else 'N/A'}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True)
    parser.add_argument("--data",  required=True, help="JSON from fetch_and_analyze")
    parser.add_argument("--sheet", choices=["portfolio", "ai", "both"], default="both",
                        help="Which sheet(s) to update: portfolio=Sheet1 only, "
                             "ai=Sheet2 only, both=default (updates both)")
    args = parser.parse_args()

    print(f"📝 Updating Excel: {args.excel}")

    # Load JSON data
    with open(args.data, encoding="utf-8") as f:
        data = json.load(f)

    stocks     = data.get("stocks", {})
    usdthb     = data.get("usdthb", 33.0)
    updated_at = data.get("updated_at", TODAY_SHORT)
    week_trade = data.get("week_trade")

    # Load Excel
    wb = openpyxl.load_workbook(args.excel)

    # Update sheets — respecting the --sheet scope argument
    if args.sheet in ("portfolio", "both"):
        # Try both sheet names for compatibility
        port_sheet_name = None
        for name in ("Portfolio", "พอร์ตหุ้นรายตัว"):
            if name in wb.sheetnames:
                port_sheet_name = name
                break
        if port_sheet_name:
            ws_port = wb[port_sheet_name]
            update_portfolio_sheet(ws_port, stocks, usdthb, updated_at)
            update_portfolio_summary(ws_port, stocks, usdthb, updated_at)
        else:
            print("   ⚠️  Portfolio sheet not found — skipped")

    if args.sheet in ("ai", "both"):
        if "AI Trading" in wb.sheetnames:
            update_ai_trading_sheet(wb["AI Trading"], stocks, week_trade, updated_at)
        else:
            print("   ⚠️  Sheet 'AI Trading' not found — skipped")

    # Save
    wb.save(args.excel)
    print(f"   💾 Excel saved successfully")


if __name__ == "__main__":
    main()
